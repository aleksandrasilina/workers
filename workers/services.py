from openpyxl.reader.excel import load_workbook


def excel_to_json_worker(filename):
    """Конвертирует данные работников из excel в json."""

    workbook = load_workbook(filename=filename, read_only=True)
    worksheet = workbook.active

    workers = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        worker = dict()
        worker["worker_id"] = int(row[0])
        worker["name"] = row[1].strip()
        worker["team_id"] = int(row[2])
        worker["salary"] = int(row[3])
        worker["specialization"] = row[4].strip()
        workers.append(worker)

    return workers
